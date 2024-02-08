import math
import os
import re
import string
import random
import openpyxl
import datetime
import pdfkit
import smtplib
import hashlib
from sqlalchemy.orm import Session
from starlette.responses import RedirectResponse
from common.consts import SMTP_SERVER_ADDR, SMTP_SERVER_FROM, SMTP_SERVER_PORT, SMTP_SERVER_PW
from jinja2 import Template
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from cryptography.fernet import Fernet  # symmetric encryption
from database.schema import Users, WhoResult


def makePageNo(pcur, pcntpg, ptotcnt):
    pg = {}
    pcur = pcur
    pcntpg = pcntpg
    ptotcnt = ptotcnt

    if (ptotcnt <= 1):  # 1 page
        pg['page_previous'] = ''# first page disable
        pg['page_next'] = ''    # next page disable
        pg['page_list'] = [1]  # page list
        pcur = 1  # current page number
        return pg

    if (ptotcnt < pcur):
        pcur = ptotcnt

    page_start_no = (math.floor((pcur-1)/pcntpg) * pcntpg) + 1
    if (page_start_no < 1):
        page_start_no = 1

    page_end_no = page_start_no + pcntpg - 1
    if (page_end_no > ptotcnt):
        page_end_no = ptotcnt

    if page_start_no > 1:
        pg['page_previous'] = page_start_no - 1  # first page disable
    else:
        pg['page_previous'] = ''   # first page disable

    pg['page_list'] = []
    for i in range(page_start_no, page_end_no+1):
        pg['page_list'].append(i)

    if (page_end_no < ptotcnt):
        pg['page_next'] = page_end_no + 1  # next page enable
    else:
        pg['page_next'] = ''    # next page disable

    # print(pcur, pg)
    return pg



def sortInvert(orderStr: str, compStr: str):
    if orderStr == compStr:
        sr = '' if (orderStr[:1] == '-') else '-'
    else:
        if orderStr[:1] == '-':
            sr = '' if orderStr[1:] == compStr else '-'
        else:
            sr = '-'

    return sr + compStr


def checkLogin(cookie):
    # 0=Logout or No user, 1=User, 2=Master / user table idd
    rt = {'level': 0, 'userid': -1}

    if type(cookie) is dict:
        emk = cookie.get('DsUsInfoKey')
        tmk = cookie.get('DsTmInfoKey')

        if (emk != None) and (tmk != None):
            em = EncToData(emk)
            tm = EncToData(tmk)

            ser = Users.filter(email=em, status='active')
            if ser:
                user = ser.first()
                if user:
                    odtm = strtimeToDatetime(tm)
                    odif = datetime.datetime.now()-odtm
                    if (odif.days == 0) and (odif.seconds >= 0) and (odif.seconds < 43200):  # login timeout(12hours)
                        rt['userid'] = user.id
                        if user.level == 'user':
                            rt['level'] = 1
                        elif user.level == 'master':
                            rt['level'] = 2
                        else:   #  error
                            rt['level'] = 0

    return rt


def getDbUserId(cookie):
    rt = -1

    if type(cookie) is dict:
        emk = cookie.get('DsUsInfoKey')
        tmk = cookie.get('DsTmInfoKey')

        if (emk != None) and (tmk != None):
            em = EncToData(emk)
            tm = EncToData(tmk)

            ser = Users.filter(email=em)
            if ser:
                user = ser.first()
                if user:
                    rt = user.id
    return rt


def reportToPdf(sampleid, form):
    pdfpath = os.getcwd()+'/static/pdf/'
    
    if os.path.exists(pdfpath+'{0}.pdf'.format(sampleid)):
        os.remove(pdfpath+'{0}.pdf'.format(sampleid))
    
    f = open(pdfpath+'report.html', 'r', encoding='utf-8', errors="ignore")
    s = f.read()
    f.close()

    tmp = Template(s)
    ret = tmp.render(form.__dict__)

    options = {'margin-top': '11', 'margin-right': '15', 'margin-bottom': '9', 'margin-left': '15', 'encoding': "UTF-8", }
    pdfkit_config = pdfkit.configuration(wkhtmltopdf=bytes('/usr/local/bin/wkhtmltopdf', 'utf-8'))
    pdfkit.from_string(ret, pdfpath+'{0}.pdf'.format(sampleid),
                       options=options, configuration=pdfkit_config)

# step code decoder
def stepnoToStr(stepno):
    if stepno == 0: return 'Not started'
    elif stepno == 1: return 'Ready'
    elif stepno == 2: return 'Ready to retry'
    elif stepno == 5: return 'Stop'
    elif stepno == 6: return 'Error'
    elif stepno == 10: return 'Running-1/5'
    elif stepno == 20: return 'Running-2/5'
    elif stepno == 30: return 'Running-3/5'
    elif stepno == 40: return 'Running-4/5'
    elif stepno == 50: return 'Running-5/5'
    elif stepno == 99: return 'Finished'
    else: return ''

def convert_size(size_bytes):
    import math
    if size_bytes == 0:
        return "0B"
    size_name = ("B", "KB", "MB", "GB", "TB", "PB", "EB", "ZB", "YB")
    i = int(math.floor(math.log(size_bytes, 1024)))
    p = math.pow(1024, i)
    s = round(size_bytes / p, 2)
    return "%s %s" % (s, size_name[i])


def ensure_range(val, minval, maxval):
    if (val <= minval):
        return minval
    elif (val >= maxval):
        return maxval
    else:
        return val

def in_range(val, minval, maxval):
    mn = min(minval, maxval)
    mx = max(minval, maxval)
    if((val>=mn)and(val<=mx)):
        return True
    else:
        return False   


def RedirectMainpage():
    response = RedirectResponse(url="/", status_code=302)
    response.set_cookie("DsUsInfoKey", None)
    response.set_cookie("DsTmInfoKey", None)
    return response


def strToDate(sdt):  # String to Date, ex) 2022-11-31 -> '2022-11-31 00:00:00'
    if type(sdt) is datetime.datetime:
        return sdt

    if (len(sdt.split('-')) > 2):
        sp = '-'
    elif (len(sdt.split('/')) > 2):
        sp = '/'
    else:
        return datetime.strptime('20000101', '%Y%m%d')
    try:
        # datetime format return
        a = int(sdt.split(sp)[0])
        b = int(sdt.split(sp)[1])
        c = int(sdt.split(sp)[2])

        if a > 31:       # 2022-11-31 -> '2022-11-31 00:00:00'
            return datetime.datetime(a, b, c)
        elif c > 31:     # 11-31-2022 -> '2022-11-31 00:00:00'
            return datetime.datetime(c, a, b)
        else:
            return datetime.strptime('20000101', '%Y%m%d')

    except Exception as e:
        print(e)
        return datetime.strptime('20000101', '%Y%m%d')


def strToInt(sd, d=0):
    try:
        rt = int(sd)
    except:
        rt = d
    return rt

def strToStr(sd, d=''):
    try:
        rt = str(sd)
    except:
        rt = d
    return rt

def strtimeToDatetime(ival):
    try:
        rt = datetime.datetime.strptime(ival, '%Y%m%d%H%M%S')
    except:
        rt = datetime.datetime.strptime('20000101000000', '%Y%m%d%H%M%S')

    return rt


def DataToEnc(email):  # Encrypt
    EnDec = SimpleEnDecrypt()
    encMail = EnDec.encrypt(email)
    return encMail


def EncToData(encMail):  # Decrypt
    denc = SimpleEnDecrypt()
    try:
        DecMail = denc.decrypt(encMail)
    except:
        return ''

    return DecMail


# String to Email, DateTime, 2022/1/2 03:04:05 ex) 20220102030405abc@abc.com
def KeyToRepasswordEmail(keys):
    denc = SimpleEnDecrypt()
    dkeys = denc.decrypt(keys)
    if (len(dkeys) < 20):
        return '', ''
    else:
        otm = strtimeToDatetime(dkeys[0:14])
        oem = dkeys[14:]

    if otm == datetime.datetime.strptime('20000101000000', '%Y%m%d%H%M%S'):
        return '', otm

    return oem, otm


# String to Email, DateTime, 2022/1/2 03:04:05 ex) 20220102030405abc@abc.com
def RepasswordEmailToKey(email):
    EnDec = SimpleEnDecrypt()
    code = EnDec.encrypt(
        datetime.datetime.now().strftime('%Y%m%d%H%M%S') + email)
    return code


def md5make(instr):
    pk = 'MYCO#^MYCO'+instr
    enc = hashlib.md5()
    enc.update(pk.encode())
    return enc.hexdigest()


def ConfirmEmail(Email: str, Session: Session):
    if Email == '':
        return False

    # User ID(email)
    ser = Users.filter(session=Session,  email=Email)
    if ser and ser.first():
        return True
    else:
        return False


def ConfirmEncEmail(EncEmail: str, Session: Session):
    if EncEmail == '':
        return False

    Email = EncToData(EncEmail)
    return ConfirmEmail(Email, Session)


def excelDataLoad(fname):  # Sample batch file loading
    print('exccel load : ', fname)
    if not os.path.exists(fname):
        return []

    # column Index position
    try:
        Data = []
        idx = {}
        cnt = 0

        xlsx = openpyxl.load_workbook(fname, data_only=True)
        sheet = xlsx.worksheets[0]      # Sheet name : Annotation
        xlsx.close
        for col in sheet['1']:      # title
            if col.value == 'Sample ID':
                idx['Sample_ID'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'R1 Filename':
                idx['R1_Filename'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'R2 Filename':
                idx['R2_Filename'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'Patient ID':
                idx['Patient_ID'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'Patient Name':
                idx['Patient_Name'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'Sample Recevied Date':
                idx['Sample_Recevied_Date'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'Library Prep Date':
                idx['Library_Prep_Date'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'Sequencing Date':
                idx['Sequencing_Date'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'Laboratory Technician':
                idx['Laboratory_Technician'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'Contact':
                idx['Contact'] = col.column-1
                cnt = cnt + 1

        if (cnt != 10):  # Header check
            return Data

    except Exception as e:
        print(e)
        return Data

    # data loading
    i = 0
    for row in sheet.rows:
        if i > 0:
            sid = row[idx['Sample_ID']].value
            if sid and sid != '':
                try:
                    dt = {}
                    dt['Analysis_R1_Data_Id'] = row[idx['R1_Filename']].value
                    dt['Analysis_R2_Data_Id'] = row[idx['R2_Filename']].value
                    dt['Sample_DataID'] = row[idx['Sample_ID']].value
                    dt['Sample_PatientID'] = row[idx['Patient_ID']].value
                    dt['Sample_PatientName'] = row[idx['Patient_Name']].value
                    dt['Sample_RecvDate'] = strToDate(
                        row[idx['Sample_Recevied_Date']].value)
                    dt['Sample_LibPrepDate'] = strToDate(
                        row[idx['Library_Prep_Date']].value)
                    dt['Sample_SeqDate'] = strToDate(
                        row[idx['Sequencing_Date']].value)
                    dt['Sample_ReportDate'] = datetime.datetime.strptime(
                        '20000101', '%Y%m%d')
                    dt['Sample_LabTechnician'] = row[idx['Laboratory_Technician']].value
                    dt['Sample_Contact'] = row[idx['Contact']].value
                    dt['Analysis_Step1Time'] = datetime.datetime.strptime(
                        '20000101', '%Y%m%d')
                    dt['Analysis_Step2Time'] = datetime.datetime.strptime(
                        '20000101', '%Y%m%d')
                    dt['Analysis_Step3Time'] = datetime.datetime.strptime(
                        '20000101', '%Y%m%d')
                    dt['Analysis_Step4Time'] = datetime.datetime.strptime(
                        '20000101', '%Y%m%d')

                    Data.append(dt)
                except Exception as e:
                    print(e)
        i = i+1

    return Data


def addDbFileLoad(fname, DBname, usid, session: Session):  # WHO DB batch file loading
    print('exccel load : ', fname)
    if not os.path.exists(fname):
        return False
       
    # column Index position
    try:
        cnt = 0
        idx = {}

        xlsx = openpyxl.load_workbook(fname, data_only=True)
        sheet = xlsx.worksheets[0]      # Sheet name : Annotation
        xlsx.close
        for col in sheet['1']:      # title
            if col.value == 'Target_Gene':
                idx['Target_Gene'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'Accession_ID':
                idx['Accession_ID'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'Start':
                idx['Start'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'End':
                idx['End'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'Ref':
                idx['Ref'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'Alt':
                idx['Alt'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'Variant_Type':
                idx['Variant_Type'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'HGVS':
                idx['HGVS'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'HGVS_AA':
                idx['HGVS_AA'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'Tier':
                idx['Tier'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'AMK':
                idx['AMK'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'BDQ':
                idx['BDQ'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'CAP':
                idx['CAP'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'CFZ':
                idx['CFZ'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'DLM':
                idx['DLM'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'EMB':
                idx['EMB'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'ETO':
                idx['ETO'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'INH':
                idx['INH'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'KAN':
                idx['KAN'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'LFX':
                idx['LFX'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'LZD':
                idx['LZD'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'MFX':
                idx['MFX'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'PZA':
                idx['PZA'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'RIF':
                idx['RIF'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'STM':
                idx['STM'] = col.column-1
                cnt = cnt + 1
            elif col.value == 'Order_level':
                idx['Order_level'] = col.column-1
                cnt = cnt + 1
            '''
            elif col.value == 'Note':
                idx['Note'] = col.column-1
                cnt = cnt + 1
            
            elif col.value == 'References':
                idx['References'] = col.column-1
                cnt = cnt + 1
            '''

        if (cnt != 26):  # Header check
            return False

    except Exception as e:
        print(e)
        return False

    # data loading
    headck = False
    for row in sheet.rows:
        if headck:
            sid = row[idx['Target_Gene']].value
            if sid and sid != '':
                try:
                    dt = {}
                    dt['Gene'] = strToStr(row[idx['Target_Gene']].value)
                    dt['GeneID'] = strToStr(row[idx['Accession_ID']].value)
                    dt['StartPs'] = strToInt(row[idx['Start']].value)
                    dt['EndPs'] = strToInt(row[idx['End']].value)
                    dt['Ref'] = strToStr(row[idx['Ref']].value)
                    dt['Alt'] = strToStr(row[idx['Alt']].value)
                    dt['VarType'] = strToStr(row[idx['Variant_Type']].value)
                    dt['Necleotide'] = strToStr(row[idx['HGVS']].value)
                    dt['AminoAcid'] = strToStr(row[idx['HGVS_AA']].value)
                    dt['Tier'] = strToInt(row[idx['Tier']].value)
                    dt['AMK'] = strToInt(row[idx['AMK']].value)
                    dt['BDQ'] = strToInt(row[idx['BDQ']].value)
                    dt['CAP'] = strToInt(row[idx['CAP']].value)
                    dt['CFZ'] = strToInt(row[idx['CFZ']].value)
                    dt['DLM'] = strToInt(row[idx['DLM']].value)
                    dt['EMB'] = strToInt(row[idx['EMB']].value)
                    dt['ETO'] = strToInt(row[idx['ETO']].value)
                    dt['INH'] = strToInt(row[idx['INH']].value)
                    dt['KAN'] = strToInt(row[idx['KAN']].value)
                    dt['LFX'] = strToInt(row[idx['LFX']].value)
                    dt['LZD'] = strToInt(row[idx['LZD']].value)
                    dt['MFX'] = strToInt(row[idx['MFX']].value)
                    dt['PZA'] = strToInt(row[idx['PZA']].value)
                    dt['RIF'] = strToInt(row[idx['RIF']].value)
                    dt['STM'] = strToInt(row[idx['STM']].value)
                    dt['Orderlev'] = strToInt(row[idx['Order_level']].value)
                    '''
                    dt['Note'] = strToStr(row[idx['Note']].value)
                    dt['Reference'] = strToStr(row[idx['References']].value)
                    '''
                    dt['user_id'] = usid
                    dt['db_name'] = DBname
                    
                    WhoResult.create(session=session, **dt)
                except Exception as e:
                    print(e)
        headck = True
    session.commit()

    return True



def saveExcelFile(fname, sData):  # Sample list excel save

    # delte file 
    if os.path.isfile(fname):
        os.remove(fname)
        
    try:
        wb = openpyxl.Workbook()
        sheet = wb.worksheets[0]     

        for rno, rowlt in enumerate(sData): 
            clen = len(rowlt)
            
            for cno, colitem in enumerate(rowlt):
                sheet.cell(row=rno+1, column=cno+1, value=colitem)
                
        for cno in sheet.columns:
            clen = max(len(str(cell.value))*1.1 for cell in cno)
            sheet.column_dimensions[cno[0].column_letter].width = clen
            
        wb.save(fname)
        return True
    except Exception as e:
        print(e)
        return False



def generatorPW():
    pw_candidate = string.ascii_letters+string.digits+'_-^*#'
    pw = ''
    for i in range(10):
        pw += random.choice(pw_candidate)
    return pw


def verifyPw(pw:str):
    em = re.compile('^[A-Za-z0-9`~!@#\$%\^&\*\(\)\{\}\[\]\_\+\|;:<>,\./\?]{8,30}$')
    if em.match(pw):
        return True
    else:
        return False

def verifyEmail(email:str):
    em = re.compile('^[a-zA-Z0-9+-_.]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$')
    if em.match(email):
        return True
    else:
        return False

class SimpleEnDecrypt:
    def __init__(self, key=None):
        self.key = b'aWHAgB-j9xfdtdAJDOah3aGOJ5MnPPt7FuyBA9dDWic='  #KEY
        self.f   = Fernet(self.key)
    
    def encrypt(self, data, is_out_string=True):
        if isinstance(data, bytes):
            ou = self.f.encrypt(data) 
        else:
            ou = self.f.encrypt(data.encode('utf-8')) # encoding
        if is_out_string is True:
            return ou.decode('utf-8') # decoding 
        else:
            return ou
        
    def decrypt(self, data, is_out_string=True):
        if isinstance(data, bytes):
            ou = self.f.decrypt(data)
        else:
            ou = self.f.decrypt(data.encode('utf-8')) # encoding
        if is_out_string is True:
            return ou.decode('utf-8') # decoding
        else:
            return ou
    

def sendActivateEmail(toEmail): 
    
    code = RepasswordEmailToKey(toEmail)
    
    body = ['Dear User',
        'To activate your account, visit the following link within log in with your code:',
        '',
        'http://www.mycochase.org/user/activate?key={0}'.format(code),
        '',
        'Thank you very much',
        'GenoMycAnalyzer Service','',
        'This email and any attachments are intended for the sole use of the named recipient(s) and contain(s) confidential information that may be proprietary, privileged or copyrighted under applicable law.',
        'If you are not the intended recipient, do not read, copy, or forward this email message or any attachments.'
        'Delete this email message and any attachments immediately.',
        'This is an automatically generated E-Mail. Do not reply to this message. Any reply to this address will not be read.']

    msg = MIMEText('\n'.join(body))
    msg['Subject'] = 'To activate your account'
    msg['From'] = SMTP_SERVER_FROM
    msg['To'] = toEmail

    smtp = smtplib.SMTP_SSL(SMTP_SERVER_ADDR, 465)
    smtp.login(SMTP_SERVER_FROM, SMTP_SERVER_PW)
    smtp.sendmail(SMTP_SERVER_FROM, toEmail, msg.as_string())
    smtp.quit()
    
    

def sendRepaswordEmail(toEmail):
    code = RepasswordEmailToKey(toEmail)
                
    body = ['Dear User',
        'To reset your password, click the address link below.',
        '',
        'http://www.mycochase.org/user/repassword?Repwkey={0}'.format(code),
        '',
        'Thank you very much',
        'GenoMycAnalyzer Service','',
        'This email and any attachments are intended for the sole use of the named recipient(s) and contain(s) confidential information that may be proprietary, privileged or copyrighted under applicable law.',
        'If you are not the intended recipient, do not read, copy, or forward this email message or any attachments.',
        'Delete this email message and any attachments immediately.',
        'This is an automatically generated E-Mail. Do not reply to this message. Any reply to this address will not be read.']

    msg = MIMEText('\n'.join(body))
    msg['Subject'] = 'To activate your account'
    msg['From'] = SMTP_SERVER_FROM
    msg['To'] = toEmail

    smtp = smtplib.SMTP_SSL(SMTP_SERVER_ADDR, 465)
    smtp.login(SMTP_SERVER_FROM, SMTP_SERVER_PW)
    smtp.sendmail(SMTP_SERVER_FROM, toEmail, msg.as_string())
    smtp.quit()
    
    
